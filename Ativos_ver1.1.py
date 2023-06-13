import yfinance as yf
import datetime as dt
import openpyxl

# Define a data do dia anterior
today = dt.datetime.today()
yesterday = today - dt.timedelta(days=1)

# Obtém o preço de fechamento do dólar comercial do dia anterior
dolar = yf.download('USDBRL=X', start=yesterday, end=yesterday)
dolar_fechamento = round(dolar['Close'][0], 2)

# Obtém o preço atual do Bitcoin em REAL
bitcoin = yf.download('BTC-USD', start=yesterday, end=today)
bitcoin_fechamento = round(bitcoin['Close'][-1] * dolar_fechamento, 2)


# Abre a planilha com os ativos
workbook = openpyxl.load_workbook('ativos.xlsx')
worksheet = workbook.active

# Insere o valor do fechamento do dólar comercial na décima coluna e segunda linha
worksheet.cell(2, 10, value=dolar_fechamento)

# Insere o valor do preço do Bitcoin em dólar na décima PRIMEIRA coluna e SEGUNDA linha
worksheet.cell(2, 11, value=bitcoin_fechamento)

# SALVA AS PARADAS
workbook.save('ativos.xlsx')

# Formata a data de execução
date_str = today.strftime('%d/%m/%Y')

# Abre a planilha com os ativos
workbook = openpyxl.load_workbook('ativos.xlsx')
worksheet = workbook.active

# Adiciona uma nova coluna para armazenar os preços de fechamento
worksheet.cell(1, 2, value='Preço de Fechamento')
worksheet.cell(1, 3, value='P/L')

# Obtém o preço de fechamento e o P/L do dia anterior para cada ativo
precos = []
pls = []
for row in worksheet.iter_rows(min_row=2, values_only=True):
    ativo = str(row[0])
    if ativo and not ativo.isspace():  # verifica se a célula não é vazia ou só contém espaços
        data = yf.download(ativo, start=yesterday, end=today)
        if not data.empty:
            preco_fechamento = round(data['Close'][0], 2)
            precos.append(preco_fechamento)
            p_l = preco_fechamento / data['Adj Close'][0] # Calcula o P/L
            pls.append(p_l)
        else:
            precos.append(None)
            pls.append(None)
    else:
        precos.append(None)
        pls.append(None)

# Atualiza a planilha com os preços de fechamento e P/L obtidos
for i, row in enumerate(worksheet.iter_rows(min_row=2)):
    preco_fechamento = precos[i]
    pl = pls[i]
    if preco_fechamento is not None:
        row[1].value = preco_fechamento
    if pl is not None:
        row[2].value = pl

# Define o formato da célula como número com duas casas decimais
for col in worksheet.columns:
    for cell in col:
        if isinstance(cell.value, float):
            cell.number_format = '0.00'

# Exibe o conteúdo da planilha
for row in worksheet.iter_rows(values_only=True):
    print(row)

# Salva a planilha com os preços atualizados
workbook.save('ativos.xlsx')
